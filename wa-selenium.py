from os import name, sep, times
from openpyxl.cell import cell
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import csv
import time
import pandas as pd
import openpyxl


def element_presence(by, xpath, time):
    element_present = EC.presence_of_element_located((By.XPATH, xpath))
    WebDriverWait(driver, time).until(element_present)


def send_message(receiver, message):
    element_presence(
        By.XPATH, '//*[@id="side"]/div[1]/div/label/div/div[2]', 30)
    receiver_element = driver.find_element(
        By.XPATH, '//*[@id="side"]/div[1]/div/label/div/div[2]')
    receiver_element.send_keys(receiver)
    receiver_element.send_keys("\n")
    time.sleep(1)
    element_presence(
        By.XPATH, '//*[@id="main"]//footer//div[contains(@contenteditable, "true")]', 30)
    msg_box = driver.find_element(
        By.XPATH, '//*[@id="main"]//footer//div[contains(@contenteditable, "true")]')
    msg_box.send_keys(message)
    msg_box.send_keys('\n')


chrome_options = Options()

#chromedriver path location
driver = webdriver.Chrome('chromedriver')
# driver = webdriver.Chrome(options=chrome_options)
driver.get("https://web.whatsapp.com/")

message = "Selamat pagi bapak/ibu <receiver>, pesan ini dikirim otomatis dari sistem untuk mengingatkan.... Terima Kasih "


# Contact file patch
path = "test-contact.xlsx"

# workbook object is created
wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active
max_col = sheet_obj.max_column
max_row = sheet_obj.max_row

# Loop will print all columns name, max row can be edited to be maximum contact list
for i in range(1, max_row + 1):
    cell_obj1 = sheet_obj.cell(row=i, column=1)
    cell_obj2 = sheet_obj.cell(row=i, column=2)
    number = str(cell_obj1.value)
    name = str(cell_obj2.value)

    # print(number, name)
    print(number, message.replace("<receiver>", name))
    send_message(number, message.replace("<receiver>", name))
